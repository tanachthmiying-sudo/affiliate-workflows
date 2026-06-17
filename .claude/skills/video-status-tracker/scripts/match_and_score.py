#!/usr/bin/env python3
"""
video-status-tracker helper: parse a TikTok Seller export, match it against
Creator's Pool rows pulled from Feishu Base, score growth potential, and
compute AFS conversion rates.

This script does NOT talk to Feishu itself — lark-cli (run from the skill
instructions) handles all reads/writes. This script only crunches the two
JSON-able inputs (TikTok export + Creator's Pool rows) into the values that
get written back.

Usage as a library (preferred — import and call from a one-off Python
snippet in the skill flow):

    from match_and_score import (
        load_tiktok_export, normalize_creator_id, classify_growth_potential,
        match_pool_to_videos, compute_afs_conversion_rates,
    )

Usage as a CLI (handy for a quick standalone check):

    python match_and_score.py --tiktok-file export.xlsx --dump-columns
        -> prints detected column names, to sanity-check auto-detection

See SKILL.md for the full step-by-step workflow this plugs into.
"""

import argparse
import json
import re
import sys
from collections import defaultdict
from datetime import datetime


# ── Column auto-detection ────────────────────────────────────────────────
# TikTok Seller exports rename/reorder columns across versions and locales.
# Match by substring/keyword rather than exact name so small drifts don't
# break the workflow. Order matters: more specific patterns first.
COLUMN_ALIASES = {
    "title": ["video name", "video title"],
    "link": ["video link", "video url"],
    "post_date": ["video post date", "post date", "publish date"],
    "creator": ["creator username", "creator id", "author username", "creator"],
    "gmv": ["gmv"],  # NOTE: exports often also have "Affiliate shoppable video GMV" —
                      # that's a different, usually larger, number (broader attribution
                      # window). Default to the plain "GMV" column; if a TikTok export
                      # only has the "shoppable video GMV" variant, ask the user which
                      # one they mean before assuming — it changes Growth Potential math.
    "orders": ["affiliate orders", "orders"],
    "views": ["shoppable video impressions", "video views", "views", "impressions"],
    "likes": ["shoppable video likes", "likes"],
    "comments": ["shoppable video comments", "comments"],
    "shares": ["shoppable video shares", "shares"],
}


def detect_columns(header_row):
    """Map our canonical field names to actual column names found in the file.

    Returns (mapping, missing) where mapping is {canonical: actual_column_name}
    and missing is a list of canonical names with no match — surface these to
    the user instead of silently leaving the metric blank or guessing.
    """
    header_lower = {str(h).strip().lower(): h for h in header_row if h}
    mapping = {}
    missing = []
    for canonical, aliases in COLUMN_ALIASES.items():
        found = None
        for alias in aliases:
            for h_lower, h_actual in header_lower.items():
                if alias == h_lower or alias in h_lower:
                    found = h_actual
                    break
            if found:
                break
        if found:
            mapping[canonical] = found
        else:
            missing.append(canonical)
    return mapping, missing


def normalize_creator_id(raw):
    """Case-insensitive, strip '@' prefix and whitespace — the matching key
    between TikTok's 'Creator username' and Feishu's 'Creators ID' field."""
    if not raw:
        return ""
    s = str(raw).strip().lower()
    if s.startswith("@"):
        s = s[1:]
    return s.strip()


def to_float(x):
    """TikTok exports use '--' for zero/empty and thousands separators."""
    if x is None:
        return 0.0
    s = str(x).strip()
    if s in ("", "--", "-"):
        return 0.0
    s = s.replace(",", "").replace("%", "")
    try:
        return float(s)
    except ValueError:
        return 0.0


def to_int(x):
    return int(round(to_float(x)))


def parse_date(value):
    """TikTok export dates may come through as datetime objects (xlsx) or
    'YYYY-MM-DD' strings (csv/xlsx-as-text). Returns a datetime or None."""
    if isinstance(value, datetime):
        return value
    if not value:
        return None
    s = str(value).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(s[: len(fmt) + 2], fmt)
        except ValueError:
            continue
    return None


def load_tiktok_export(path, column_map):
    """Load an xlsx or csv TikTok export into a list of normalized video dicts.

    column_map comes from detect_columns() — pass it in rather than
    re-detecting here so the skill can show the user the detected mapping
    once and reuse it.
    """
    rows = []
    if path.lower().endswith((".xlsx", ".xlsm")):
        import openpyxl

        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))
        header, data_rows = all_rows[0], all_rows[1:]
    else:
        import csv

        with open(path, newline="", encoding="utf-8-sig") as f:
            reader = list(csv.reader(f))
        header, data_rows = reader[0], reader[1:]

    hidx = {name: i for i, name in enumerate(header)}

    def get(row, canonical):
        col = column_map.get(canonical)
        if col is None or col not in hidx:
            return None
        return row[hidx[col]]

    for row in data_rows:
        if not row or get(row, "creator") is None:
            continue
        link = get(row, "link") or ""
        m = re.search(r"/video/(\d+)", str(link))
        video_id = m.group(1) if m else ""
        rows.append(
            {
                "creator_norm": normalize_creator_id(get(row, "creator")),
                "creator_raw": get(row, "creator"),
                "video_id": video_id,
                "title": str(get(row, "title") or "").strip(),
                "post_date": parse_date(get(row, "post_date")),
                "views": to_int(get(row, "views")) if "views" in column_map else None,
                "likes": to_int(get(row, "likes")) if "likes" in column_map else None,
                "comments": to_int(get(row, "comments")) if "comments" in column_map else None,
                "shares": to_int(get(row, "shares")) if "shares" in column_map else None,
                "orders": to_int(get(row, "orders")) if "orders" in column_map else 0,
                "gmv": to_float(get(row, "gmv")) if "gmv" in column_map else 0.0,
            }
        )
    return rows


def classify_growth_potential(views, gmv, orders):
    """🔥 High / ⚡ Medium / ⬇ Low. None-safe — missing metrics count as 0."""
    views = views or 0
    gmv = gmv or 0
    orders = orders or 0
    if views > 50000 or gmv > 5000 or orders > 50:
        return "🔥 High"
    if views > 10000 or gmv > 1000 or orders > 10:
        return "⚡ Medium"
    return "⬇ Low"


def match_pool_to_videos(pool_rows, videos, start, end):
    """pool_rows: list of {record_id, creator_id_raw, afs_owner, date_of_contact}
    videos: list from load_tiktok_export
    start/end: datetime bounds (inclusive) for the video post date

    Returns (matched, not_posted). For creators with multiple videos in the
    window, picks the highest-GMV one — that's the one worth surfacing for
    Boost Ads, and avoids double-counting the same creator.
    """
    by_creator = defaultdict(list)
    for v in videos:
        if v["post_date"] and start <= v["post_date"] <= end:
            by_creator[v["creator_norm"]].append(v)

    matched, not_posted = [], []
    for p in pool_rows:
        cid = normalize_creator_id(p.get("creator_id_raw"))
        vids = by_creator.get(cid) if cid else None
        if not vids:
            not_posted.append(p)
            continue
        best = max(vids, key=lambda v: v["gmv"])
        potential = classify_growth_potential(best["views"], best["gmv"], best["orders"])
        matched.append(
            {
                **p,
                "video": best,
                "potential": potential,
                "boost": potential == "🔥 High",
                "video_count": len(vids),
            }
        )
    return matched, not_posted


def compute_afs_conversion_rates(pool_rows, matched):
    """Group by AFS owner (first value if a row has multiple). Returns
    {owner: "posted/contacted (rate%)"} plus the raw counts for the summary."""
    contacted = defaultdict(int)
    posted = defaultdict(int)
    matched_ids = {m["record_id"] for m in matched}
    for p in pool_rows:
        owner = p.get("afs_owner") or "(Unassigned)"
        contacted[owner] += 1
        if p["record_id"] in matched_ids:
            posted[owner] += 1

    rates, stats = {}, {}
    for owner in contacted:
        c, ps = contacted[owner], posted[owner]
        rate = (ps / c * 100) if c else 0
        rates[owner] = f"{ps}/{c} ({rate:.0f}%)"
        stats[owner] = {"contacted": c, "posted": ps, "rate_pct": rate}
    return rates, stats


def _cli():
    parser = argparse.ArgumentParser(description="Inspect a TikTok export's detected columns")
    parser.add_argument("--tiktok-file", required=True)
    parser.add_argument("--dump-columns", action="store_true")
    args = parser.parse_args()

    if args.tiktok_file.lower().endswith((".xlsx", ".xlsm")):
        import openpyxl

        wb = openpyxl.load_workbook(args.tiktok_file, data_only=True)
        header = next(wb.active.iter_rows(values_only=True))
    else:
        import csv

        with open(args.tiktok_file, newline="", encoding="utf-8-sig") as f:
            header = next(csv.reader(f))

    mapping, missing = detect_columns(header)
    if args.dump_columns:
        print("Raw header:", list(header))
        print(json.dumps({"mapping": mapping, "missing": missing}, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    _cli()
